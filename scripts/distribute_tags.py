"""
distribute_tags.py
------------------
Reads tag JSON files written by push_to_cm360.py,
generates a clean formatted tag sheet per campaign,
and emails it to the requester and biddable team.

Uses SendGrid for email. Requires SENDGRID_API_KEY secret.
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
TAGS_DIR = ROOT / "tags"
OUTPUTS_DIR = ROOT / "outputs"
OUTPUTS_DIR.mkdir(exist_ok=True)

SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
FROM_EMAIL = os.environ.get("FROM_EMAIL", "tagging-pipeline@youragency.com")
BIDDABLE_EMAIL = os.environ.get("BIDDABLE_TEAM_EMAIL", "")

def build_tag_sheet(camp_data: dict) -> Path:
    """Build a clean formatted Excel tag sheet matching CM360 export format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking Ads"

    # Header block
    header_font = Font(bold=True, size=10)
    ws.merge_cells("B2:I2")
    ws["B2"] = "CONTRACT INFORMATION"
    ws["B2"].font = header_font

    info = [
        ("Advertiser", camp_data.get("advertiserName", "")),
        ("Campaign ID", camp_data.get("campaignId", "")),
        ("Campaign Name", camp_data.get("campaignName", "")),
        ("Generated", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for i, (label, value) in enumerate(info, 3):
        ws[f"B{i}"] = label
        ws[f"I{i}"] = value
        ws[f"B{i}"].font = header_font

    # Column headers
    headers = [
        "Placement ID", "Site", "Placement Name", "Dimensions",
        "Start Date", "End Date",
        "Impression Tag (image)", "Impression Tag (JavaScript)", "Click Tag"
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font_white = Font(color="FFFFFF", bold=True, size=10)
    header_row = 9

    for col, header in enumerate(headers, 2):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 35
    ws.freeze_panes = f"B{header_row + 1}"

    # Data rows
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F0F4F8")

    for row_idx, p in enumerate(camp_data["placements"], header_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        row_data = [
            p.get("placementId", ""),
            p.get("site", ""),
            p.get("placementName", ""),
            p.get("dimensions", ""),
            p.get("startDate", ""),
            p.get("endDate", ""),
            p.get("impressionTagImage", ""),
            p.get("impressionTagJs", ""),
            p.get("clickTag", ""),
        ]
        for col_idx, value in enumerate(row_data, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    col_widths = [14, 16, 55, 12, 12, 12, 80, 80, 80]
    for col_idx, width in enumerate(col_widths, 2):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Instructions sheet
    inst_ws = wb.create_sheet("Instructions")
    instructions = [
        ["CM360 Tag Sheet — Implementation Guide"],
        [""],
        ["Impression Tag (image)", "Use for Meta, Instagram, TikTok. Paste into the pixel/tracking field."],
        ["Impression Tag (JavaScript)", "Use for display/programmatic (StackAdapt, etc). Paste into the tag field."],
        ["Click Tag", "Use for all platforms as the click-through/destination URL tracker."],
        [""],
        ["[timestamp]", "Replace with a dynamically generated random number for cache busting."],
        ["${GDPR}", "Your platform will populate this automatically for EEA traffic."],
        ["${GDPR_CONSENT_755}", "Your platform will populate this automatically for EEA traffic."],
    ]
    for row in instructions:
        inst_ws.append(row)
    inst_ws.column_dimensions["A"].width = 30
    inst_ws.column_dimensions["B"].width = 70
    inst_ws["A1"].font = Font(bold=True, size=12)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    camp_name = camp_data.get("campaignName", "campaign").replace("/", "-")
    output_path = OUTPUTS_DIR / f"Tags_{camp_name}_{timestamp}.xlsx"
    wb.save(output_path)
    print(f"  ✓ Tag sheet written: {output_path.name}")
    return output_path

def send_email(to_emails: list, subject: str, body: str, attachment_path: Path):
    """Send tag sheet via SendGrid."""
    if not SENDGRID_API_KEY:
        print("  ⚠  SENDGRID_API_KEY not set — skipping email, tag sheet saved to outputs/")
        return

    import urllib.request
    import base64

    with open(attachment_path, "rb") as f:
        attachment_data = base64.b64encode(f.read()).decode()

    payload = {
        "personalizations": [{"to": [{"email": e} for e in to_emails]}],
        "from": {"email": FROM_EMAIL, "name": "CM360 Tagging Pipeline"},
        "subject": subject,
        "content": [{"type": "text/html", "value": body}],
        "attachments": [{
            "content": attachment_data,
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": attachment_path.name,
        }]
    }

    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=json.dumps(payload).encode(),
        headers={
            "Authorization": f"Bearer {SENDGRID_API_KEY}",
            "Content-Type": "application/json"
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req) as resp:
            print(f"  ✓ Email sent to: {', '.join(to_emails)} (status {resp.status})")
    except Exception as e:
        print(f"  ✗ Email failed: {e}")

def distribute():
    tag_files = list(TAGS_DIR.glob("tags_*.json"))
    if not tag_files:
        print("No tag files found.")
        sys.exit(0)

    for tag_file in tag_files:
        camp_data = json.loads(tag_file.read_text())
        camp_name = camp_data.get("campaignName", "Campaign")
        client = camp_data.get("client", "")
        n_placements = len(camp_data.get("placements", []))
        requester_email = camp_data.get("requesterEmail", "")

        print(f"\n→ Distributing tags for: {camp_name} ({n_placements} placements)")

        # Build tag sheet
        sheet_path = build_tag_sheet(camp_data)

        # Build email
        placement_rows = "".join([
            f"<tr><td style='padding:6px 12px;border-bottom:1px solid #eee'>{p['placementName']}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee'>{p['site']}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee'>{p['dimensions']}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee'>{p['startDate']} → {p['endDate']}</td></tr>"
            for p in camp_data["placements"]
        ])

        body = f"""
        <div style='font-family:sans-serif;max-width:600px;margin:0 auto'>
          <div style='background:#1F4E79;padding:24px;border-radius:8px 8px 0 0'>
            <h1 style='color:#fff;margin:0;font-size:20px'>CM360 Tags Ready</h1>
            <p style='color:#a8c7e8;margin:8px 0 0'>{client} — {camp_name}</p>
          </div>
          <div style='background:#fff;border:1px solid #e0e0e0;border-top:none;padding:24px;border-radius:0 0 8px 8px'>
            <p style='color:#333'>Your CM360 tags are ready for implementation. The tag sheet is attached.</p>
            <table style='width:100%;border-collapse:collapse;margin:16px 0;font-size:13px'>
              <thead>
                <tr style='background:#f5f5f5'>
                  <th style='padding:8px 12px;text-align:left;border-bottom:2px solid #ddd'>Placement</th>
                  <th style='padding:8px 12px;text-align:left;border-bottom:2px solid #ddd'>Site</th>
                  <th style='padding:8px 12px;text-align:left;border-bottom:2px solid #ddd'>Dimensions</th>
                  <th style='padding:8px 12px;text-align:left;border-bottom:2px solid #ddd'>Dates</th>
                </tr>
              </thead>
              <tbody>{placement_rows}</tbody>
            </table>
            <p style='color:#666;font-size:12px'>
              See the <strong>Instructions</strong> tab in the attached sheet for platform-specific guidance
              on which tag type to use and how to handle [timestamp] and GDPR placeholders.
            </p>
          </div>
        </div>
        """

        # Send to requester + biddable team
        recipients = list(filter(None, [requester_email, BIDDABLE_EMAIL]))
        if recipients:
            send_email(
                to_emails=recipients,
                subject=f"CM360 Tags Ready — {client} {camp_name}",
                body=body,
                attachment_path=sheet_path
            )
        else:
            print("  ⚠  No recipient emails configured — tag sheet saved to outputs/ only")

        # Archive the tag file so it's not resent
        tag_file.rename(tag_file.with_suffix(".sent.json"))

if __name__ == "__main__":
    distribute()
